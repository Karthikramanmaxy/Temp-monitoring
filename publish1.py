from awscrt import io, mqtt, auth, http
from awsiot import mqtt_connection_builder
import time as t
import json
import random
from openpyxl import *

# Define ENDPOINT, CLIENT_ID, PATH_TO_CERTIFICATE, PATH_TO_PRIVATE_KEY, PATH_TO_AMAZON_ROOT_CA_1, MESSAGE, TOPIC, and RANGE
ENDPOINT = "a7xush1fkupwu-ats.iot.us-east-1.amazonaws.com"
CLIENT_ID = "testDevice"
PATH_TO_CERTIFICATE = "certificates/a9432b6ee06d7713925178e83c14998748854a4596439e7f15a7348966c1ec60-certificate.pem.crt"
PATH_TO_PRIVATE_KEY = "certificates/a9432b6ee06d7713925178e83c14998748854a4596439e7f15a7348966c1ec60-private.pem.key"
PATH_TO_AMAZON_ROOT_CA_1 = "certificates/root.pem"
TOPIC = "device/34/data"
RANGE = 1
wb = load_workbook('C:\\Users\\User\\Desktop\\MQTT AWS\\iot-test-publish\\excel.xlsx')
sheet = wb.active
# Spin up resources
event_loop_group = io.EventLoopGroup(1)
host_resolver = io.DefaultHostResolver(event_loop_group)
client_bootstrap = io.ClientBootstrap(event_loop_group, host_resolver)
mqtt_connection = mqtt_connection_builder.mtls_from_path(
            endpoint=ENDPOINT,
            cert_filepath=PATH_TO_CERTIFICATE,
            pri_key_filepath=PATH_TO_PRIVATE_KEY,
            client_bootstrap=client_bootstrap,
            ca_filepath=PATH_TO_AMAZON_ROOT_CA_1,
            client_id=CLIENT_ID,
            clean_session=False,
            keep_alive_secs=6
            )
print("Connecting to {} with client ID '{}'...".format(
        ENDPOINT, CLIENT_ID))
# Make the connect() call
connect_future = mqtt_connection.connect()
# Future.result() waits until a result is available
connect_future.result()
print("Connected!")
# Publish message to server desired number of times.
print('Begin Publish')


for i in range(RANGE):
    row = 11
    data = {
        "temperature": sheet.cell(row+1,column=1).value,
        "humidity": sheet.cell(row+1, column=2).value,
        "barometer": sheet.cell(row+1, column=3).value,
        "wind": {
            "velocity": sheet.cell(row+1, column=4).value,
            "bearing": sheet.cell(row+1, column=5).value
        }
    };
    #message = {"message" : data}
    mqtt_connection.publish(topic=TOPIC, payload=json.dumps(data), qos=mqtt.QoS.AT_LEAST_ONCE)
    print("Published: '" + json.dumps(data["temperature"]) + "' to the topic: " + TOPIC)

print('Publish End')
disconnect_future = mqtt_connection.disconnect()
disconnect_future.result()
