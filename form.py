from openpyxl import *
from tkinter import *
 
# globally declare wb and sheet variable
 
# opening the existing excel file
wb = load_workbook('C:\\Users\\User\\Desktop\\MQTT AWS\\iot-test-publish\\excel.xlsx')
 
# create the sheet object
sheet = wb.active
 
 
def excel():
     
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
 
    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Temperature"
    sheet.cell(row=1, column=2).value = "Humidity"
    sheet.cell(row=1, column=3).value = "Barometer"
    sheet.cell(row=1, column=4).value = "Wind velocity"
    sheet.cell(row=1, column=5).value = "Wind bearing"
 
 
# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    Temperature_field.focus_set()
 
 
# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    Humidity_field.focus_set()
 
 
# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    Barometer_field.focus_set()
 
 
# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    Wind_velocity_field.focus_set()
 
 
# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    Wind_bearing_field.focus_set()
 
# Function for clearing the
# contents of text entry boxes
def clear():
     
    # clear the content of text entry box
    Temperature_field.delete(0, END)
    Humidity_field.delete(0, END)
    Barometer_field.delete(0, END)
    Wind_velocity_field.delete(0, END)
    Wind_bearing_field.delete(0, END) 
 
# Function to take data from GUI
# window and write to an excel file
def insert():
     
    # if user not fill any entry
    # then print "empty input"
    if (Temperature_field.get() == "" and 
        Humidity_field.get() == "" and
        Barometer_field.get() == "" and
        Wind_velocity_field.get() == "" and
        Wind_bearing_field.get() == ""):   
        print("empty input")
 
    else:
 
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column
 
        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = Temperature_field.get()
        sheet.cell(row=current_row + 1, column=2).value = Humidity_field.get()
        sheet.cell(row=current_row + 1, column=3).value = Barometer_field.get()
        sheet.cell(row=current_row + 1, column=4).value = Wind_velocity_field.get()
        sheet.cell(row=current_row + 1, column=5).value = Wind_bearing_field.get()
 
        # save the file
        wb.save('C:\\Users\\User\\Desktop\\MQTT AWS\\iot-test-publish\\excel.xlsx')
 
        # set focus on the name_field box
        Temperature_field.focus_set()
 
        # call the clear() function
        clear()
 
 
# Driver code
if __name__ == "__main__":
     
    # create a GUI window
    root = Tk()
 
    # set the background colour of GUI window
    root.configure(background='light green')
 
    # set the title of GUI window
    root.title("registration form")
 
    # set the configuration of GUI window
    root.geometry("500x300")
 
    excel()
 
    # create a Form label
    heading = Label(root, text="Form", bg="light green")
 
    # create a Name label
    Temperature = Label(root, text="Temperature", bg="light green")
 
    # create a Course label
    Humidity = Label(root, text="Humidity", bg="light green")
 
    # create a Semester label
    Barometer= Label(root, text="Barometer", bg="light green")
 
    # create a Form No. label
    Wind_velocity = Label(root, text="Wind velocity", bg="light green")
 
    # create a Contact No. label
    Wind_bearing= Label(root, text="Wind Bearing", bg="light green")
 
    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    Temperature.grid(row=1, column=0)
    Humidity.grid(row=2, column=0)
    Barometer.grid(row=3, column=0)
    Wind_velocity.grid(row=4, column=0)
    Wind_bearing.grid(row=5, column=0)
     
    # create a text entry box
    # for typing the information
    Temperature_field = Entry(root)
    Humidity_field = Entry(root)
    Barometer_field = Entry(root)
    Wind_velocity_field = Entry(root)
    Wind_bearing_field = Entry(root)
     
    # bind method of widget is used for
    # the binding the function with the events
 
    # whenever the enter key is pressed
    # then call the focus1 function
    Temperature_field.bind("<Return>", focus1)
 
    # whenever the enter key is pressed
    # then call the focus2 function
    Humidity_field.bind("<Return>", focus2)
 
    # whenever the enter key is pressed
    # then call the focus3 function
    Barometer_field.bind("<Return>", focus3)
 
    # whenever the enter key is pressed
    # then call the focus4 function
    Wind_velocity_field.bind("<Return>", focus4)
 
    # whenever the enter key is pressed
    # then call the focus5 function
    Wind_bearing_field.bind("<Return>", focus5)
 
    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    Temperature_field.grid(row=1, column=1, ipadx="100")
    Humidity_field.grid(row=2, column=1, ipadx="100")
    Barometer_field.grid(row=3, column=1, ipadx="100")
    Wind_velocity_field.grid(row=4, column=1, ipadx="100")
    Wind_bearing_field.grid(row=5, column=1, ipadx="100")
  
    # call excel function
    excel()
 
    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                            bg="Red", command=insert)
    submit.grid(row=6, column=1)
 
    # start the GUI
    root.mainloop()


